from pathlib import Path
from os import system, listdir
import openpyxl as xl


def merge_2_excel_files():
    file1 = excel_file_name_input(
        '1st', title='***Merge 2 Excel Files into 1st File***\n')
    file2 = excel_file_name_input(
        '2nd', title='***Merge 2 Excel Files into 1st File***\n')
    print(f'Filling  In  - {file1}\nFilling From - {file2}\nOpening Files...')

    workbook1 = xl.load_workbook(Path(file1))
    workbook2 = xl.load_workbook(Path(file2))
    sheet1 = workbook1.active
    sheet2 = workbook2.active

    max_rows1, max_cols1 = sheet1.max_row, sheet1.max_column
    max_rows2, max_cols2 = sheet2.max_row, sheet2.max_column

    if max_rows1 != max_rows2 or max_cols1 != max_cols2:
        print(f"\nFile1 -> Rows x Columns: {max_rows1} x {max_cols1}")
        print(f"File2 -> Rows x Columns: {max_rows2} x {max_cols2}")
        input("\n**Both Files doesn't have equal Rows and Columns.\n  Please Check the files.\n\n  Closing Program!!")
        exit()
    else:
        print(f"\nRows x Columns: {max_rows2} x {max_cols2}")

    flag = False
    cells_with_different_data = []
    for i in range(1, max_rows1 + 1):
        for j in range(1, max_cols1 + 1):
            if sheet1.cell(i, j).value == sheet2.cell(i, j).value:
                pass
            elif sheet1.cell(i, j).value != None and sheet2.cell(i, j).value == None:
                pass
            elif sheet1.cell(i, j).value == None and sheet2.cell(i, j).value != None:
                sheet1.cell(i, j).value = sheet2.cell(i, j).value
            elif sheet1.cell(i, j).value != None and sheet2.cell(i, j).value != None and sheet2.cell(i, j).value != sheet1.cell(i, j).value:
                flag = True
                cells_with_different_data.append(num_to_col_letter(j) + str(i))

    if flag == True:
        print(f"\n\n***There were Cells with different data in both files.")
        print("   So, Value in 1st File or Sheet is Considered.\n")
        print(cells_with_different_data)

    print("\n\n*** Merged Successfully!!! ***\n\nNew File created as 'Merge Result File.xlsx'\n\nSaving File...")
    workbook1.save("./Merge Result File.xlsx")
    input("\n*** File Saved !!! ***")


def excel_file_name_input(string='', title=''):
    system('cls')
    if string != '':
        string += ' '
    print(title)
    print("Excel files in same directory:")
    files = listdir('.')
    xl_files = []
    counter = 0
    for file in files:
        if file[-4:] == 'xlsx':
            counter += 1
            xl_files.append(file)
            print(f"{counter}. {file}")

    if counter == 0:
        input("\n\n**No Files !!!\n  Closing Program !!")
        exit()
    elif counter == 1:
        input("\n\n**Only 1 File is Present !!!\n  Closing Program !!")
        exit()
    option = int(
        input(f"\nEnter the file number corresponding to {string}file name: ")) - 1
    if len(xl_files) > option >= 0:
        system('cls')
        print(title)
        return xl_files[option]
    else:
        input("\n\n**Wrong Input !!!\n  Closing Program !!")
        exit()


def num_to_col_letter(num):
    letters = ''
    while num:
        mod = (num - 1) % 26
        letters += chr(mod + 65)
        num = (num - 1) // 26
    return ''.join(reversed(letters))

merge_2_excel_files()
